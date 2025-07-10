import os
import sqlite3
from datetime import datetime
from openpyxl import load_workbook

class ExcelImporter:
    def __init__(self, db_path):
        self.db_path = db_path
        
    def import_excel(self, file_path):
        """
        Importa datos de un archivo Excel a la base de datos SQLite
        
        Args:
            file_path (str): Ruta al archivo Excel a importar
            
        Returns:
            tuple: (total_registros, exitosos, errores)
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"El archivo {file_path} no existe")
            
        # Cargar el archivo Excel
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active
        
        # Obtener los encabezados
        headers = [cell.value for cell in sheet[1]]
        
        # Mapeo de columnas del Excel a los campos de la base de datos
        column_mapping = {
            'FECHA': 'task_date',
            'CLIENTE': 'client_name',
            'PEDIDO': 'order_number',
            'TAREA': 'task_description',
            'PRESUPUESTO': 'budget_total',
            'MANO DE OBRA': 'labor_cost',
            'MATERIAL': 'material_cost',
            'SEGURO': 'insurance_payment',
            'EFECTIVO': 'cash_payment',
            'GASTO MATERIAL': 'material_expense',
            'IVA': 'iva',
            'GANANCIA': 'profit',
            'PABLO (70%)': 'pablo_share',
            'FACU (30%)': 'facu_share'
        }
        
        # Verificar que todas las columnas necesarias estén presentes
        missing_columns = [col for col in column_mapping.keys() if col not in headers]
        if missing_columns:
            raise ValueError(f"Faltan columnas requeridas en el Excel: {', '.join(missing_columns)}")
        
        # Conectar a la base de datos
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        total = 0
        success = 0
        errors = []
        
        # Procesar cada fila (empezando desde la fila 2 para omitir encabezados)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Crear diccionario con los datos de la fila
                row_data = {column_mapping[header]: value 
                           for header, value in zip(headers, row) 
                           if header in column_mapping}
                
                # Convertir fecha de string a formato SQLite (YYYY-MM-DD)
                if 'task_date' in row_data and row_data['task_date']:
                    if isinstance(row_data['task_date'], str):
                        # Si es string, convertir a datetime
                        date_obj = datetime.strptime(row_data['task_date'], '%d/%m/%Y')
                    else:
                        # Si ya es un objeto date
                        date_obj = row_data['task_date']
                    row_data['task_date'] = date_obj.strftime('%Y-%m-%d')
                
                # Reemplazar comas por puntos en los números
                for key, value in row_data.items():
                    if isinstance(value, str):
                        # Si es un número con coma decimal
                        if ',' in value and value.replace(',', '').replace('.', '').isdigit():
                            # Reemplazar comas por puntos y convertir a float
                            row_data[key] = float(value.replace('.', '').replace(',', '.'))
                
                # Insertar en la base de datos
                columns = ', '.join(row_data.keys())
                placeholders = ', '.join(['?'] * len(row_data))
                values = tuple(row_data.values())
                
                # Asegurarse de que la tabla existe
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS tasks (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        task_date TEXT,
                        client_name TEXT,
                        order_number TEXT,
                        task_description TEXT,
                        budget_total REAL,
                        labor_cost REAL,
                        material_cost REAL,
                        insurance_payment REAL,
                        cash_payment REAL,
                        material_expense REAL,
                        iva REAL,
                        profit REAL,
                        pablo_share REAL,
                        facu_share REAL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Insertar el registro
                cursor.execute(
                    f"INSERT INTO tasks ({columns}) VALUES ({placeholders})",
                    values
                )
                
                success += 1
            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")
            finally:
                total += 1
        
        # Guardar cambios y cerrar conexión
        conn.commit()
        conn.close()
        
        return total, success, errors


def import_excel_file():
    """
    Muestra un diálogo para seleccionar e importar un archivo Excel.
    
    Returns:
        bool: True si la importación fue exitosa, False en caso contrario
    """
    from PySide6.QtWidgets import QFileDialog, QMessageBox
    
    # Abrir diálogo para seleccionar archivo
    file_path, _ = QFileDialog.getOpenFileName(
        None,
        "Seleccionar archivo Excel",
        "",
        "Archivos Excel (*.xlsx *.xls);;Todos los archivos (*)"
    )
    
    if not file_path:
        return
    
    try:
        # Crear instancia del importador
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
        importer = ExcelImporter(db_path)
        
        # Importar datos
        total, success, errors = importer.import_excel(file_path)
        
        # Mostrar resultados
        msg_parts = [
            "Importación completada.",
            "",
            f"Registros procesados: {total}",
            f"Registros importados: {success}",
            f"Errores: {len(errors)}"
        ]
        
        if errors:
            msg_parts.extend([
                "",
                "Errores encontrados:",
                *[str(e) for e in errors[:10]]  # Mostrar solo los primeros 10 errores
            ])
            if len(errors) > 10:
                msg_parts.append(f"... y {len(errors) - 10} errores más.")
                
        msg = "\n".join(msg_parts)
        
        QMessageBox.information(
            None,
            "Importación completada",
            msg,
            QMessageBox.StandardButton.Ok
        )
        
        return True
    
    except Exception as e:
        QMessageBox.critical(
            None,
            "Error en la importación",
            f"Ocurrió un error al importar el archivo:\n{str(e)}",
            QMessageBox.StandardButton.Ok
        )
        return False
