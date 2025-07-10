import os
import re
import pandas as pd
import json
from datetime import datetime, timedelta
from .task_dialog import TaskDialog
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
    QTableWidget, QTableWidgetItem, QHeaderView, 
    QComboBox, QLabel, QMessageBox, QDateEdit, QFileDialog,
    QGroupBox, QGridLayout, QTabWidget, QFrame, QSizePolicy
)
from PySide6.QtGui import QColor
from PySide6.QtCore import Qt, QDate, Signal
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class ReportView(QWidget):
    # Señal personalizada para mensajes de estado
    status_message = Signal(str)
    
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setup_styles()
        self.init_ui()
        
    def setup_styles(self):
        """Configura los estilos para los widgets de la interfaz con un tema oscuro."""
        # Colores principales
        self.dark_bg = "#1e1e2f"
        self.darker_bg = "#161623"
        self.card_bg = "#27293d"
        self.text_color = "#e9ecef"
        self.accent_color = "#2196F3"  # Azul del logo
        self.accent_hover = "#1976D2"   # Azul más oscuro para hover
        self.secondary_color = "#2196F3"  # Mismo azul para consistencia
        self.border_color = "#2b3553"
        self.highlight_color = "#3a3f5a"
        
        # Estilo para los botones principales
        self.button_style = f"""
            QPushButton {{
                background-color: {self.accent_color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
                min-width: 100px;
                text-transform: uppercase;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {self.accent_hover};
            }}
            QPushButton:pressed {{
                background-color: #0d47a1;
                padding-top: 9px;
                padding-bottom: 7px;
            }}
            QPushButton:disabled {{
                background-color: #90caf9;
                color: #e3f2fd;
            }}
        """
        
        # Aplicar estilos generales
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {self.dark_bg};
                color: {self.text_color};
            }}
            QLabel {{
                color: {self.accent_color};
                font-weight: 500;
            }}
            QGroupBox {{
                border: 1px solid {self.border_color};
                border-radius: 6px;
                margin-top: 10px;
                padding: 10px;
                font-weight: bold;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
                color: {self.accent_color};
            }}
            QTableWidget {{
                background-color: {self.card_bg};
                border: 1px solid {self.border_color};
                border-radius: 6px;
                gridline-color: {self.border_color};
                font-size: 11px;
            }}
            QTableWidget::item {{
                padding: 5px;
                border-bottom: 1px solid {self.border_color};
            }}
            QTableWidget::item:selected {{
                background-color: {self.highlight_color};
                color: white;
            }}
            QHeaderView::section {{
                background-color: {self.darker_bg};
                color: {self.accent_color};
                padding: 8px;
                border: none;
                font-weight: 600;
                font-size: 11px;
            }}
            QComboBox, QDateEdit {{
                background-color: {self.darker_bg};
                color: {self.text_color};
                border: 1px solid {self.border_color};
                border-radius: 4px;
                padding: 5px;
                min-width: 120px;
            }}
            QComboBox::drop-down {{
                border-left: 1px solid {self.border_color};
                width: 20px;
            }}
            QComboBox QAbstractItemView {{
                background-color: {self.darker_bg};
                color: {self.text_color};
                selection-background-color: {self.accent_color};
                selection-color: white;
            }}
        """)
    
    def init_ui(self):
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(5, 5, 5, 5)  # Reducir márgenes del layout principal
        main_layout.setSpacing(5)  # Reducir espaciado entre widgets
        
        # Filtros
        filter_group = QGroupBox("Filtros")
        filter_layout = QHBoxLayout()
        
        # Selección de técnico
        self.technician_combo = QComboBox()
        self.technician_combo.setMinimumWidth(250)
        
        # Fechas
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_edit.setCalendarPopup(True)
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        
        # Layout principal para los filtros
        filter_layout = QHBoxLayout()
        
        # Agregar widgets de filtro
        filter_layout.addWidget(QLabel("Técnico:"))
        filter_layout.addWidget(self.technician_combo)
        filter_layout.addWidget(QLabel("Desde:"))
        filter_layout.addWidget(self.start_date_edit)
        filter_layout.addWidget(QLabel("Hasta:"))
        filter_layout.addWidget(self.end_date_edit)
        
        # Botón de búsqueda
        search_btn = QPushButton("🔍 Buscar")
        search_btn.clicked.connect(self.load_report)
        search_btn.setStyleSheet(self.button_style)
        search_btn.setCursor(Qt.PointingHandCursor)
        filter_layout.addWidget(search_btn)
        
        filter_layout.addStretch()
        filter_group.setLayout(filter_layout)
        
        # Layout para los botones de acción
        action_group = QGroupBox("Acciones")
        action_layout = QHBoxLayout()
        
        # Grupo de botones de tareas
        tasks_group = QGroupBox("Tareas")
        tasks_layout = QHBoxLayout()
        
        self.add_button = QPushButton("➕ Agregar Tarea")
        self.edit_button = QPushButton("✏️ Editar Tarea")
        self.delete_button = QPushButton("🗑️ Eliminar Tarea")
        
        # Aplicar estilos a los botones
        for btn in [self.add_button, self.edit_button, self.delete_button]:
            btn.setStyleSheet(self.button_style)
            btn.setCursor(Qt.PointingHandCursor)
        
        tasks_layout.addWidget(self.add_button)
        tasks_layout.addWidget(self.edit_button)
        tasks_layout.addWidget(self.delete_button)
        tasks_group.setLayout(tasks_layout)
        
        # Grupo de botones de importar/exportar
        io_group = QGroupBox("Importar/Exportar")
        io_layout = QHBoxLayout()
        
        self.export_button = QPushButton("📤 Exportar Datos")
        self.export_button.setCursor(Qt.PointingHandCursor)
        self.import_button = QPushButton("📥 Importar Datos")
        self.import_button.setCursor(Qt.PointingHandCursor)
        self.template_button = QPushButton("📋 Descargar Plantilla")
        self.template_button.setCursor(Qt.PointingHandCursor)
        
        # Aplicar estilos a los botones de importar/exportar
        for btn in [self.export_button, self.import_button, self.template_button]:
            btn.setStyleSheet(self.button_style)
        
        io_layout.addWidget(self.export_button)
        io_layout.addWidget(self.import_button)
        io_layout.addWidget(self.template_button)
        io_group.setLayout(io_layout)
        
        # Agregar grupos al layout principal de botones
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)  # Eliminar márgenes internos
        button_layout.setSpacing(10)  # Reducir espacio entre widgets
        
        # Crear un widget contenedor para los grupos
        container = QWidget()
        container_layout = QHBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(20)  # Espacio entre grupos
        
        container_layout.addWidget(tasks_group)
        container_layout.addWidget(io_group)
        container_layout.addStretch()  # Empuja los grupos hacia la izquierda
        
        button_layout.addWidget(container)
        action_layout.addLayout(button_layout)
        action_group.setLayout(action_layout)
        
        # Ajustar el tamaño de los grupos
        tasks_group.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Preferred)
        io_group.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Preferred)
        
        # Configurar márgenes y espaciado
        main_layout.setSpacing(2)  # Mínimo espacio entre widgets
        main_layout.setContentsMargins(4, 4, 4, 4)  # Márgenes mínimos
        
        # Agregar grupos al layout principal
        main_layout.addWidget(filter_group)
        main_layout.addWidget(action_group)
        
        # Estilo para los grupos de botones
        group_style = """
            QGroupBox {
                border: 1px solid #2b3553;
                border-radius: 4px;
                margin-top: 10px;
                padding-top: 20px;
                background: transparent;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #2196F3;
                font-weight: bold;
            }
        """
        
        # Aplicar estilos a los grupos
        filter_group.setStyleSheet(group_style)
        action_group.setStyleSheet(group_style)
        
        # Contenedor principal para el contenido de reportes
        
        # Pestaña de resumen
        self.summary_tab = QWidget()
        summary_layout = QVBoxLayout()
        
        # Resumen de métricas
        metrics_layout = QHBoxLayout()
        
        # Grupo de métricas con el mismo estilo que los demás grupos
        self.metrics_group = QGroupBox("Resumen")
        metrics_inner_layout = QHBoxLayout()
        self.metrics_group.setStyleSheet(group_style)
        
        self.total_tasks_label = QLabel("📋 Tareas: 0")
        self.total_income_label = QLabel("💰 Ingresos: $0.00")
        self.total_profit_label = QLabel("💵 Ganancia: $0.00")
        self.technician_payment_label = QLabel("👷 Pago a técnico: $0.00")
        self.material_cost_label = QLabel("🔧 Costo material: $0.00")
        
        # Aplicar estilos a las etiquetas de métricas
        for label in [self.total_tasks_label, self.total_income_label, self.total_profit_label, 
                     self.technician_payment_label, self.material_cost_label]:
            label.setStyleSheet(f"""
                font-weight: bold; 
                font-size: 12px;
                color: {self.accent_color};
                padding: 5px 10px;
                background: transparent;
                margin: 0;
            """)
            metrics_inner_layout.addWidget(label)
        
        # Agregar las etiquetas al layout
        for label in [self.total_tasks_label, self.total_income_label, self.total_profit_label,
                     self.technician_payment_label, self.material_cost_label]:
            metrics_inner_layout.addWidget(label)
        
        metrics_inner_layout.addStretch()
        self.metrics_group.setLayout(metrics_inner_layout)
        
        # Tabla de tareas
        self.tasks_table = QTableWidget()
        self.tasks_table.setColumnCount(16)  
        self.tasks_table.setHorizontalHeaderLabels([
            "ID", "Cliente", "Tarea", "Fecha", "Presupuesto Total", 
            "Mano Obra", "Material", "Pago Seguro", "Efectivo",
            "N° Pedido", "Gasto Material", "Ganancia Total", 
            "Técnico (70%)", "Seguro Técnico", "Socio (30%)", "IVA", "Estado"
        ])   
        
        # Ajustar ancho de columnas
        self.tasks_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  
        self.tasks_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  
        
        # Configurar layout del contenido
        content_layout = QVBoxLayout()
        content_layout.addWidget(self.metrics_group)
        content_layout.addWidget(self.tasks_table)
        
        # Agregar widgets al layout principal
        main_layout.addLayout(content_layout)
        
        self.setLayout(main_layout)
        
        # Conectar señales
        self.add_button.clicked.connect(self.add_task)
        self.edit_button.clicked.connect(self.edit_task)
        self.delete_button.clicked.connect(self.delete_task)
        self.export_button.clicked.connect(self.export_data)
        self.import_button.clicked.connect(self.import_data)
        self.template_button.clicked.connect(self.download_template)
        
        # Cargar técnicos
        self.load_technicians()
    
    def load_technicians(self):
        """Carga la lista de técnicos en el combo box"""
        self.technician_combo.clear()
        
        # Agregar la opción "Todos los técnicos" al inicio
        self.technician_combo.addItem("Todos los técnicos", None)
        
        # Cargar el resto de los técnicos
        technicians = self.db.get_technicians()
        for tech in technicians:
            self.technician_combo.addItem(tech['name'], tech['id'])
    
    def load_report(self):
        """Carga el reporte según los filtros seleccionados"""
        print("\n--- Cargando reporte ---")
        technician_id = self.technician_combo.currentData()
        print(f"Técnico ID: {technician_id}")
        
        # Guardar el reporte actual para exportación
        self.current_report = None
        
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().addDays(1).toString("yyyy-MM-dd")
        print(f"Rango de fechas: {start_date} - {end_date}")
        
        # Obtener datos del reporte
        if technician_id is None:  # Si se seleccionó 'Todos los técnicos'
            # Obtener todos los técnicos
            technicians = self.db.get_technicians()
            all_tasks = []
            summary = {
                'total_tasks': 0,
                'total_income': 0,
                'total_profit': 0,
                'total_insurance_payment': 0,
                'total_cash_payment': 0,
                'total_material_expense': 0,
                'total_iva': 0,
                'technician_share': 0,
                'facu_share': 0
            }
            
            # Obtener tareas para cada técnico
            print(f"Buscando tareas para {len(technicians)} técnicos...")
            for tech in technicians:
                print(f"  - Buscando tareas para técnico {tech['id']}: {tech['name']}")
                report = self.db.generate_report(tech['id'], start_date, end_date)
                if report and 'tasks' in report and report['tasks']:
                    print(f"    - Encontradas {len(report['tasks'])} tareas")
                    all_tasks.extend(report['tasks'])
                    # Sumar al resumen general
                    if 'summary' in report:
                        s = report['summary']
                        for key in summary.keys():
                            if key in s:
                                summary[key] += s[key]
                else:
                    print(f"    - No se encontraron tareas para el técnico {tech['name']}")
            
            if not all_tasks:
                QMessageBox.warning(self, "Advertencia", "No se encontraron datos para el rango seleccionado")
                return
                
            # Actualizar métricas con el resumen consolidado
            self.update_metrics(summary)
            
            # Actualizar tabla de tareas
            self.update_tasks_table(all_tasks)
            
            # Guardar el reporte actual para exportación
            self.current_report = {
                'tasks': all_tasks,
                'summary': summary
            }
        else:  # Reporte para un técnico específico
            if not technician_id:
                QMessageBox.warning(self, "Advertencia", "Por favor seleccione un técnico")
                return
                
            report = self.db.generate_report(technician_id, start_date, end_date)
            
            if not report or 'tasks' not in report:
                QMessageBox.warning(self, "Advertencia", "No se encontraron datos para el rango seleccionado")
                return
            
            # Obtener nombre del técnico para mostrarlo en el reporte
            technician_name = self.technician_combo.currentText()
            
            # Actualizar métricas
            self.update_metrics(report.get('summary', {}))
            
            # Agregar nombre del técnico a cada tarea para mostrarlo en la tabla
            for task in report['tasks']:
                task['technician_name'] = technician_name
            
            # Actualizar tabla de tareas
            self.update_tasks_table(report['tasks'])
            
            # Guardar el reporte actual para exportación
            self.current_report = report
    
    def update_metrics(self, summary):
        """Actualiza las métricas del reporte"""
        if not summary:
            return
            
        # Actualizar métricas individuales
        self.total_tasks_label.setText(f"Tareas: {summary.get('total_tasks', 0)}")
        self.total_income_label.setText(f"Ingresos: ${summary.get('total_income', 0):,.2f}")
        self.total_profit_label.setText(f"Ganancia: ${summary.get('total_profit', 0):,.2f}")
        self.technician_payment_label.setText(f"Pago a técnico: ${summary.get('technician_share', 0):,.2f}")
        self.material_cost_label.setText(f"Costo material: ${summary.get('total_material_expense', 0):,.2f}")
        
        # Mostrar resumen en la consola para depuración
        print("\n--- Resumen del Reporte ---")
        print(f"Total Tareas: {summary.get('total_tasks', 0)}")
        print(f"Ingresos Totales: ${summary.get('total_income', 0):,.2f}")
        print(f"Ganancia Total: ${summary.get('total_profit', 0):,.2f}")
        print(f"Total Pago Seguro: ${summary.get('total_insurance_payment', 0):,.2f}")
        print(f"Total Efectivo: ${summary.get('total_cash_payment', 0):,.2f}")
        print(f"Total Gasto Material: ${summary.get('total_material_expense', 0):,.2f}")
        print(f"Total IVA (10.5%): ${summary.get('total_iva', 0):,.2f}")
        print(f"Técnico (70%): ${summary.get('technician_share', 0):,.2f}")
        print(f"Facu (30%): ${summary.get('facu_share', 0):,.2f}")
        
        # Mostrar resumen semanal si hay datos
        if 'weekly_totals' in summary and summary['weekly_totals']:
            print("\n--- Resumen Semanal ---")
            for week_start, week_data in summary['weekly_totals'].items():
                week_end = week_data['week_start'] + timedelta(days=6)
                print(f"Semana {week_start} al {week_end.strftime('%Y-%m-%d')}:")
                print(f"  • Tareas: {week_data['tasks']}")
                print(f"  • Ingresos: ${week_data['income']:,.2f}")
                print(f"  • Ganancia: ${week_data['profit']:,.2f}")
                print(f"  • Técnico (70%): ${week_data['technician_share']:,.2f}")
                print(f"  • Facu (30%): ${week_data['facu_share']:,.2f}")
                print(f"  • Gasto Material: ${week_data['material_expense']:,.2f}")
        
        # Mostrar resumen mensual si hay datos
        if 'monthly_totals' in summary and summary['monthly_totals']:
            print("\n--- Resumen Mensual ---")
            for month_start, month_data in summary['monthly_totals'].items():
                print(f"Mes {month_data['month'].strftime('%B %Y')}:")
                print(f"  • Tareas: {month_data['tasks']}")
                print(f"  • Ingresos: ${month_data['income']:,.2f}")
                print(f"  • Ganancia: ${month_data['profit']:,.2f}")
                print(f"  • Técnico (70%): ${month_data['technician_share']:,.2f}")
                print(f"  • Facu (30%): ${month_data['facu_share']:,.2f}")
                print(f"  • Gasto Material: ${month_data['material_expense']:,.2f}")
        
    def update_tasks_table(self, tasks):
        """Actualiza la tabla de tareas con los datos proporcionados"""
        print(f"\n--- Actualizando tabla con {len(tasks)} tareas ---")
        self.tasks_table.setRowCount(len(tasks))
        
        for row, task in enumerate(tasks):
            # Obtener nombre del técnico
            technician_name = ""
            if 'technician_name' in task:
                technician_name = task['technician_name']
            
            # Formatear fecha
            task_date = task.get('task_date', '')
            if task_date:
                try:
                    # Si es un string, intentar convertir al formato deseado
                    if isinstance(task_date, str):
                        # Si ya está en formato DD/MM/YYYY, dejarlo como está
                        if re.match(r'\d{2}/\d{2}/\d{4}', task_date):
                            pass
                        # Si está en formato YYYY-MM-DD, convertir a DD/MM/YYYY
                        elif re.match(r'\d{4}-\d{2}-\d{2}', task_date):
                            task_date = datetime.strptime(task_date, '%Y-%m-%d').strftime('%d/%m/%Y')
                        # Si es otro formato, intentar parsear con dateutil
                        else:
                            from dateutil import parser
                            task_date = parser.parse(str(task_date)).strftime('%d/%m/%Y')
                    # Si es un objeto date o datetime, formatearlo
                    elif hasattr(task_date, 'strftime'):
                        task_date = task_date.strftime('%d/%m/%Y')
                except Exception as e:
                    print(f"Error al formatear fecha '{task_date}': {str(e)}")
                    # Si hay un error, mostrar la fecha original como texto
                    task_date = str(task_date)
            
            # Obtener valores numéricos para los cálculos
            profit = float(task.get('profit', 0) or 0)
            material_expense = float(task.get('material_expense', 0) or 0)
            insurance_payment = float(task.get('insurance_payment', 0) or 0)
            facu_share = float(task.get('facu_share', 0) or 0)
            iva = float(task.get('iva', 0) or 0)
            
            # Calcular valores derivados
            technician_share = (profit * 0.7) + material_expense
            seguro_tecnico = insurance_payment - facu_share - iva
            
            # Datos de la tarea - Ordenados según los encabezados de la tabla
            data = [
                int(task.get('id', 0)),  # ID
                task.get('client_name', ''),  # Cliente
                str(task.get('task_description', '')).strip(),  # Tarea
                task_date,  # Fecha
                task.get('budget_total', 0),  # Presupuesto Total
                task.get('labor_cost', 0),  # Mano Obra
                task.get('material_cost', 0),  # Material
                task.get('insurance_payment', 0),  # Pago Seguro
                task.get('cash_payment', 0),  # Efectivo
                str(task.get('order_number', '')),  # N° Pedido (como texto)
                material_expense,  # Gasto Material
                profit,  # Ganancia Total
                technician_share,  # Técnico (70%) = (GananciaTotal * 0.7) + GastoMaterial
                max(0, seguro_tecnico),  # Seguro Técnico = PagoSeguro - Socio (30%) - IVA (mínimo 0)
                facu_share,  # Socio (30%)
                iva,  # IVA
                task.get('status', 'PENDIENTE')  # Estado
            ]
            
            # Guardar el ID de la tarea en el item para referencia
            task_id = task.get('id')
            
            # Agregar datos a la tabla
            for col, value in enumerate(data):
                item = QTableWidgetItem(str(value) if value is not None else '')
                
                # Alinear números a la derecha y texto a la izquierda
                if col > 3:  # A partir de la columna 4 (Presupuesto Total en adelante)
                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).replace('-', '').isdigit()):
                        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        
                        # No formatear como moneda la columna de N° Pedido (columna 9)
                        if col == 9:  # N° Pedido
                            item.setText(str(value).strip() if value is not None else '')
                        else:
                            # Formatear números con separadores de miles y 2 decimales
                            try:
                                num_value = float(value) if value is not None else 0
                                item.setText(f"${num_value:,.2f}")
                                
                                # Resaltar ganancias en verde y pérdidas en rojo
                                if col in [11, 12, 13, 14]:  # Columnas numéricas que pueden ser positivas o negativas
                                    if num_value >= 0:
                                        item.setForeground(QColor('#006400'))  # Verde oscuro para positivos
                                    else:
                                        item.setForeground(QColor('#8B0000'))  # Rojo oscuro para negativos
                            except (ValueError, TypeError):
                                pass
                
                # Guardar el ID de la tarea en los datos del ítem para referencia futura
                if col == 0:  
                    item.setData(Qt.ItemDataRole.UserRole, task.get('id'))
                
                self.tasks_table.setItem(row, col, item)
        
        # Ajustar el ancho de las columnas al contenido
        self.tasks_table.resizeColumnsToContents()
        
        # Asegurar que la columna de descripción tenga un ancho mínimo
        self.tasks_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        
    def add_task(self):
        """Abre el diálogo para agregar una nueva tarea"""
        dialog = TaskDialog(self)
        if dialog.exec() == dialog.DialogCode.Accepted:
            task_data = dialog.get_task_data()
            
            try:
                task_id = self.db.add_task(task_data)
                if task_id:
                    QMessageBox.information(
                        self, "Éxito", 
                        f"Tarea agregada exitosamente con ID: {task_id}"
                    )
                    # Actualizar el reporte para el técnico seleccionado
                    self.load_report()
                else:
                    QMessageBox.critical(
                        self, "Error", 
                        "No se pudo agregar la tarea. Intente nuevamente."
                    )
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", 
                    f"Error al agregar la tarea:\n{str(e)}"
                )
    
    def edit_task(self):
        """Abre el diálogo para editar la tarea seleccionada"""
        selected_rows = self.tasks_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Error", "Debe seleccionar una tarea para editar")
            return
            
        row = selected_rows[0].row()
        task_id_item = self.tasks_table.item(row, 0)  # Obtener el ítem de la primera columna (ID)
        
        if not task_id_item:
            QMessageBox.critical(self, "Error", "No se pudo obtener el ID de la tarea")
            return
            
        try:
            task_id = int(task_id_item.text())
            print(f"Editando tarea ID: {task_id}")
            task_data = self.db.get_task(task_id)
            
            if not task_data:
                QMessageBox.critical(self, "Error", "No se encontró la tarea seleccionada en la base de datos")
                return
                
            print("Datos de la tarea a editar:", task_data)
            
            # Asegurarse de que los campos numéricos sean flotantes
            for key in ['budget_total', 'labor_cost', 'material_cost', 'insurance_payment', 
                       'cash_payment', 'material_expense', 'profit', 'technician_share', 'facu_share', 'iva']:
                if key in task_data and task_data[key] is not None:
                    task_data[key] = float(task_data[key])
                    
            dialog = TaskDialog(self, task_data)
            
        except ValueError as ve:
            QMessageBox.critical(self, "Error", f"El ID de la tarea no es válido: {ve}")
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar la tarea: {str(e)}")
            return
        if dialog.exec() == dialog.DialogCode.Accepted:
            updated_data = dialog.get_task_data()
            print("Datos actualizados del diálogo:", updated_data)
            
            try:
                if self.db.update_task(task_id, updated_data):
                    QMessageBox.information(
                        self, "Éxito", 
                        "Tarea actualizada exitosamente"
                    )
                    self.load_report()  # Recargar la vista
                else:
                    QMessageBox.critical(
                        self, "Error", 
                        "No se pudo actualizar la tarea en la base de datos."
                    )
            except Exception as e:
                error_msg = f"Error al actualizar la tarea:\n{str(e)}\n\n"
                error_msg += f"Tipo de error: {type(e).__name__}"
                print(error_msg)
                QMessageBox.critical(
                    self, "Error", 
                    error_msg
                )
    
    def delete_task(self):
        """Elimina la tarea seleccionada"""
        try:
            selected_rows = self.tasks_table.selectionModel().selectedRows()
            if not selected_rows:
                QMessageBox.warning(self, "Error", "Debe seleccionar una tarea para eliminar")
                return
                
            row = selected_rows[0].row()
            task_id_item = self.tasks_table.item(row, 0)  # Obtener el ítem de la primera columna (ID)
            client_item = self.tasks_table.item(row, 2)  # Obtener el ítem de la columna de cliente
            
            if not task_id_item or not client_item:
                QMessageBox.critical(self, "Error", "No se pudo obtener la información de la tarea")
                return
                
            try:
                task_id = int(task_id_item.text())
                client = client_item.text()
                
                reply = QMessageBox.question(
                    self, 'Confirmar eliminación',
                    f"¿Está seguro que desea eliminar la tarea de {client}?\n\n"
                    "Esta acción no se puede deshacer.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    print(f"Intentando eliminar tarea ID: {task_id}")
                    # Verificar si la tarea existe antes de intentar eliminarla
                    existing_task = self.db.get_task(task_id)
                    if not existing_task:
                        QMessageBox.warning(
                            self, "Advertencia",
                            "La tarea seleccionada ya no existe en la base de datos."
                        )
                        self.load_report()  # Actualizar la vista
                        return
                        
                    if self.db.delete_task(task_id):
                        # Guardar el estado actual de los filtros
                        current_tech_id = self.technician_combo.currentData()
                        start_date = self.start_date_edit.date().toString('yyyy-MM-dd')
                        end_date = self.end_date_edit.date().toString('yyyy-MM-dd')
                        
                        # Recargar la vista
                        self.load_report()
                        
                        # Restaurar los filtros
                        if current_tech_id:
                            index = self.technician_combo.findData(current_tech_id)
                            if index >= 0:
                                self.technician_combo.setCurrentIndex(index)
                        
                        QMessageBox.information(
                            self, "Éxito", 
                            "Tarea eliminada exitosamente"
                        )
                    else:
                        QMessageBox.critical(
                            self, "Error", 
                            "No se pudo eliminar la tarea en la base de datos. "
                            "Verifique que la tarea aún exista y tenga los permisos necesarios."
                        )
                        
            except ValueError as ve:
                QMessageBox.critical(
                    self, 
                    "Error de formato", 
                    f"El ID de la tarea no es válido: {ve}"
                )
            except Exception as e:
                error_msg = "Ocurrió un error inesperado al intentar eliminar la tarea.\n\n"
                error_msg += f"Error: {str(e)}\n"
                error_msg += f"Tipo: {type(e).__name__}"
                print(error_msg)  # Para depuración
                
                QMessageBox.critical(
                    self, 
                    "Error al eliminar", 
                    error_msg
                )
        except Exception as outer_e:
            # Capturar cualquier otro error inesperado
            error_msg = "Error inesperado al procesar la eliminación.\n"
            error_msg += f"Detalles: {str(outer_e)}"
            print(error_msg)  # Para depuración
            
            QMessageBox.critical(
                self,
                "Error del sistema",
                error_msg
            )
    
    def generate_general_report(self):
        """Genera un reporte general con las ganancias por técnico y el total de Facu"""
        try:
            # Obtener fechas seleccionadas
            start_date = self.start_date_edit.date().toString('yyyy-MM-dd')
            end_date = self.end_date_edit.date().toString('yyyy-MM-dd')
            
            # Consulta para obtener las ganancias por técnico
            query = """
                SELECT 
                    t.name as technician_name,
                    SUM(tr.technician_share) as total_technician,
                    SUM(tr.facu_share) as total_facu_share
                FROM tasks tr
                LEFT JOIN technicians t ON tr.technician_id = t.id
                WHERE date(tr.task_date) BETWEEN ? AND ?
                GROUP BY t.name
                ORDER BY total_technician DESC
            """
            
            self.db.cursor.execute(query, (start_date, end_date))
            results = self.db.cursor.fetchall()
            
            if not results:
                QMessageBox.information(self, "Sin datos", "No hay registros en el período seleccionado.")
                return
            
            # Calcular totales
            total_technician = sum(row[1] for row in results)
            total_facu = sum(row[2] for row in results)
            
            # Crear mensaje con los resultados
            report_lines = [
                f"Reporte General del {start_date} al {end_date}\n",
                "=" * 50,
                "{:<20} {:<20} {:<20}".format("TÉCNICO", "GANANCIA TÉCNICO", "GANANCIA FACU"),
                "-" * 60
            ]
            
            for row in results:
                technician, tech_earning, facu_earning = row
                report_lines.append(
                    "{:<20} ${:<19.2f} ${:<19.2f}".format(
                        technician or 'Sin técnico',
                        float(tech_earning or 0),
                        float(facu_earning or 0)
                    )
                )
            
            report_lines.extend([
                "-" * 60,
                "{:<20} ${:<19.2f} ${:<19.2f}".format(
                    "TOTALES:",
                    total_technician,
                    total_facu
                )
            ])
            
            # Mostrar reporte en un diálogo
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QDialogButtonBox
            
            dialog = QDialog(self)
            dialog.setWindowTitle("Reporte General")
            dialog.resize(600, 400)
            
            layout = QVBoxLayout(dialog)
            
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setFontFamily("Courier New")
            text_edit.setPlainText("\n".join(report_lines))
            
            buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            buttons.accepted.connect(dialog.accept)
            
            layout.addWidget(text_edit)
            layout.addWidget(buttons)
            
            # Opción para exportar a archivo de texto
            export_btn = QPushButton("Exportar a TXT")
            export_btn.clicked.connect(lambda: self.export_report_to_txt("\n".join(report_lines)))
            layout.addWidget(export_btn)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error al generar reporte",
                f"Ocurrió un error al generar el reporte:\n{str(e)}"
            )
    
    def export_report_to_txt(self, report_text):
        """Exporta el reporte a un archivo de texto"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Reporte",
            f"reporte_general_{QDate.currentDate().toString('yyyyMMdd')}.txt",
            "Archivos de Texto (*.txt);;Todos los archivos (*)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(report_text)
                QMessageBox.information(
                    self,
                    "Exportación exitosa",
                    f"El reporte se ha guardado en:\n{file_path}"
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Error al exportar",
                    f"No se pudo guardar el archivo:\n{str(e)}"
                )
    
    def import_excel(self):
        """Importa datos desde un archivo Excel"""
        from ...excel_importer import import_excel_file
        if import_excel_file():
            self.load_report()  # Recargar datos después de importar
    
    def export_data(self):
        """Muestra un diálogo para seleccionar el tipo de exportación"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QPushButton, QLabel, QButtonGroup, QMessageBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Seleccionar tipo de exportación")
        dialog.setMinimumWidth(500)
        dialog.setMinimumHeight(300)
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        title = QLabel("Seleccione el tipo de exportación:")
        title.setStyleSheet("""
            font-weight: bold; 
            font-size: 16px; 
            margin-bottom: 20px;
            color: #2c3e50;
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Estilo para los botones
        button_style = """
            QPushButton {
                text-align: left;
                padding: 12px 15px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background-color: #f8f9fa;
                color: #2c3e50;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e9ecef;
                border-color: #95a5a6;
            }
            QPushButton:pressed {
                background-color: #dee2e6;
            }
        """
        
        # Opción 1: Plantilla con datos (para importar)
        btn_template = QPushButton("1. Plantilla con datos (para importar en otro sistema)")
        btn_template.setToolTip("Exporta los datos en el mismo formato que la plantilla de importación")
        btn_template.setStyleSheet(button_style)
        btn_template.clicked.connect(lambda: [dialog.accept(), self.export_template_with_data()])
        
        # Opción 2: Datos completos para análisis
        btn_full_data = QPushButton("2. Datos completos para análisis")
        btn_full_data.setToolTip("Exporta todos los datos de las tareas para su análisis")
        btn_full_data.setStyleSheet(button_style)
        btn_full_data.clicked.connect(lambda: [dialog.accept(), self.export_complete_data()])
        
        # Opción 3: Reporte general por técnico
        btn_tech_report = QPushButton("3. Reporte general por técnico")
        btn_tech_report.setToolTip("Exporta un resumen de ganancias por técnico")
        btn_tech_report.setStyleSheet(button_style)
        btn_tech_report.clicked.connect(lambda: [dialog.accept(), self.export_technician_report()])
        
        # Opción 4: Reporte de Facu (30%)
        btn_facu_report = QPushButton("4. Reporte de participación de Facu (30%)")
        btn_facu_report.setToolTip("Exporta el reporte de la participación de Facu (30%)")
        btn_facu_report.setStyleSheet(button_style)
        btn_facu_report.clicked.connect(lambda: [dialog.accept(), self.export_facu_report()])
        
        # Agregar botones al layout
        for btn in [btn_template, btn_full_data, btn_tech_report, btn_facu_report]:
            btn.setMinimumHeight(50)
            layout.addWidget(btn)
        
        # Espaciador
        layout.addStretch()
        
        # Botón de cancelar
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setStyleSheet("""
            QPushButton {
                padding: 10px;
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_cancel.clicked.connect(dialog.reject)
        layout.addWidget(btn_cancel)
        
        dialog.setLayout(layout)
        
        # Mostrar el diálogo modal
        dialog.exec()
    
    def import_data(self):
        """Importa tareas desde un archivo Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar Tareas",
            "",
            "Archivos Excel (*.xlsx);;Todos los archivos (*)"
        )
        
        if not file_path:
            return
            
        try:
            # Leer el archivo Excel ignorando validaciones
            xls = pd.ExcelFile(file_path, engine='openpyxl')
            
            # Verificar que exista la hoja de Tareas
            if 'Tareas' not in xls.sheet_names:
                QMessageBox.critical(
                    self,
                    "Error",
                    "El archivo no contiene una hoja llamada 'Tareas'"
                )
                return
                
            # Leer la hoja de Tareas, convirtiendo todo a string para evitar problemas de validación
            df = pd.read_excel(xls, sheet_name='Tareas', dtype=str)
            
            # Reemplazar valores NaN con cadenas vacías
            df = df.fillna('')
            
            # Verificar columnas requeridas
            required_columns = [
                'Técnico', 'Cliente', 'Tarea',
                'Presupuesto Total', 'Efectivo', 'Pago Seguro'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Faltan columnas requeridas en el archivo: {', '.join(missing_columns)}"
                )
                return
            
            # Mostrar resumen de importación
            total_tasks = len(df)
            reply = QMessageBox.question(
                self,
                "Confirmar Importación",
                f"Se van a importar {total_tasks} tareas.\n"
                "¿Desea continuar?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Obtener lista de técnicos para validación (usando nombres en minúsculas para comparación insensible a mayúsculas)
            technicians = self.db.get_technicians()
            tech_name_to_id = {tech['name'].lower().strip(): tech['id'] for tech in technicians}
            
            # Procesar cada tarea
            success_count = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    # Obtener el nombre del técnico
                    tech_name = str(row['Técnico']).strip() if pd.notna(row['Técnico']) else ''
                    
                    if not tech_name:
                        errors.append(f"Fila {idx+2}: El campo 'Técnico' está vacío")
                        continue
                    
                    # Limpiar el nombre (eliminar cualquier texto entre paréntesis como IDs)
                    tech_name_clean = re.sub(r'\s*\(.*?\)', '', tech_name).strip()
                    tech_name_lower = tech_name_clean.lower()
                    
                    # Buscar el técnico por nombre (insensible a mayúsculas)
                    if tech_name_lower in tech_name_to_id:
                        technician_id = tech_name_to_id[tech_name_lower]
                    else:
                        # Si no se encuentra, intentar buscar coincidencias parciales
                        matching_techs = [tech for tech in technicians if tech_name_lower in tech['name'].lower()]
                        
                        if len(matching_techs) == 1:
                            # Si hay exactamente una coincidencia, usarla
                            technician_id = matching_techs[0]['id']
                        elif len(matching_techs) > 1:
                            # Si hay múltiples coincidencias, mostrar un error
                            error_msg = f"Fila {idx+2}: Múltiples técnicos coinciden con '{tech_name_clean}'. Por favor, sea más específico."
                            errors.append(error_msg)
                            continue
                        else:
                            # Si no hay coincidencias, mostrar un error
                            errors.append(f"Fila {idx+2}: Técnico no encontrado: '{tech_name_clean}'. Técnicos disponibles: {', '.join(tech_name_to_id.keys())}")
                            continue
                    
                    # Crear diccionario con los datos de la tarea
                    task_data = {
                        'technician_id': technician_id,
                        'client_name': str(row['Cliente']).strip(),
                        'task_description': str(row.get('Tarea', '')).strip(),
                        'task_date': row.get('Fecha (AAAA-MM-DD)', datetime.now().strftime('%Y-%m-%d')),
                        'budget_total': float(row.get('Presupuesto Total', 0)),
                        'labor_cost': float(row.get('Mano de Obra', 0)),
                        'material_cost': float(row.get('Presupuesto Materiales', 0)),
                        'insurance_payment': float(row.get('Pago Seguro', 0)),
                        'cash_payment': float(row.get('Efectivo', 0)),
                        'material_expense': float(row.get('Gasto Material', 0)),
                        'payment_type': str(row.get('Tipo de Pago', 'EFECTIVO')).strip().upper(),
                        'order_number': str(row.get('Número Pedido', '')).strip(),
                        'status': str(row.get('Estado', 'PENDIENTE')).strip().upper()
                    }
                    
                    # Validar fechas
                    try:
                        date_str = str(task_data['task_date']).strip()
                        
                        if not date_str or date_str.lower() in ['nan', 'nat', 'none', '']:
                            # Si la fecha está vacía, usar la fecha actual
                            task_data['task_date'] = datetime.now().strftime('%Y-%m-%d')
                            continue
                            
                        # Primero intentar con el formato de Excel (mm/dd/yyyy)
                        try:
                            # Intentar parsear como fecha de Excel (mm/dd/yyyy)
                            date_parsed = datetime.strptime(date_str, '%m/%d/%Y')
                            # Convertir al formato deseado (yyyy-mm-dd para la base de datos)
                            task_data['task_date'] = date_parsed.strftime('%Y-%m-%d')
                            continue  # Si tuvo éxito, pasar a la siguiente tarea
                        except (ValueError, TypeError):
                            # Si falla, continuar con otros formatos
                            pass
                        
                        # Si no es formato de Excel, intentar otros formatos comunes
                        date_formats = [
                            '%d/%m/%Y',  # Formato dd/mm/yyyy
                            '%d-%m-%Y',  # Formato dd-mm-yyyy
                            '%Y-%m-%d',  # Formato ISO yyyy-mm-dd
                            '%d.%m.%Y',  # Formato dd.mm.yyyy
                            '%d/%m/%y',  # Formato dd/mm/yy (año de 2 dígitos)
                            '%d-%m-%y'   # Formato dd-mm-yy (año de 2 dígitos)
                        ]
                        
                        date_parsed = None
                        
                        # Intentar con los demás formatos
                        for date_format in date_formats:
                            try:
                                date_parsed = datetime.strptime(date_str, date_format)
                                # Si llegamos aquí, el formato es correcto
                                break
                            except (ValueError, TypeError):
                                continue
                        
                        # Si no se pudo parsear, intentar con pandas
                        if date_parsed is None:
                            try:
                                date_parsed = pd.to_datetime(date_str, dayfirst=True)
                            except (ValueError, TypeError):
                                pass
                        
                        # Si aún no se pudo parsear, usar la fecha actual
                        if date_parsed is None:
                            errors.append(f"Fila {idx+2}: Formato de fecha no reconocido: '{date_str}'. Se usará la fecha actual.")
                            date_parsed = datetime.now()
                        
                        # Convertir al formato estándar de la aplicación (yyyy-mm-dd)
                        task_data['task_date'] = date_parsed.strftime('%Y-%m-%d')
                        
                    except Exception as e:
                        errors.append(f"Fila {idx+2}: Error al procesar la fecha '{task_data['task_date']}': {str(e)}. Se usará la fecha actual.")
                        task_data['task_date'] = datetime.now().strftime('%Y-%m-%d')
                    
                    # Validar estado
                    if task_data['status'] not in ['PENDIENTE', 'COMPLETADA']:
                        task_data['status'] = 'PENDIENTE'
                    
                    # Validar tipo de pago
                    if task_data['payment_type'] not in ['EFECTIVO', 'TRANSFERENCIA']:
                        task_data['payment_type'] = 'EFECTIVO'
                    
                    # Insertar la tarea
                    if self.db.add_task(task_data):
                        success_count += 1
                    else:
                        errors.append(f"Fila {idx+2}: Error al guardar la tarea")
                        
                except Exception as e:
                    errors.append(f"Fila {idx+2}: Error - {str(e)}")
            
            # Mostrar resumen
            msg = f"Se importaron {success_count} de {total_tasks} tareas correctamente."
            
            if errors:
                msg += "\n\nErrores:\n" + "\n".join(errors[:10])  # Mostrar solo los primeros 10 errores
                if len(errors) > 10:
                    msg += f"\n... y {len(errors) - 10} errores más."
            
            QMessageBox.information(
                self,
                "Importación completada",
                msg
            )
            
            # Recargar la vista
            self.load_report()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al importar el archivo: {str(e)}"
            )
    
    def download_template(self):
        """Versión simplificada de la función de descarga de plantilla"""
        try:
            # 1. Obtener ruta para guardar el archivo
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Plantilla de Tareas",
                "tareas_importar.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if not file_path:
                return
                
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            # 2. Obtener lista de técnicos
            technicians = self.db.get_technicians()
            if not technicians:
                raise ValueError("No hay técnicos registrados en el sistema")
            
            # 3. Crear un nuevo libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Tareas"
            
            # 4. Agregar encabezados
            headers = [
                "Técnico", "Cliente", "Tarea", "Presupuesto Total",
                "Mano de Obra", "Presupuesto Materiales", "Pago Seguro",
                "Efectivo", "Número Pedido", "Gasto Material",
                "Tipo de Pago", "Estado", "Fecha (AAAA-MM-DD)"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
            
            # 5. Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 20  # Técnico
            ws.column_dimensions['B'].width = 25  # Cliente
            ws.column_dimensions['C'].width = 40  # Tarea
            
            # 6. Agregar validación para técnicos
            # Primero, crear una hoja oculta con la lista de técnicos
            ws_tech = wb.create_sheet("Tecnicos")
            ws_tech.sheet_state = "hidden"
            
            # Escribir la lista de técnicos en la hoja oculta
            for i, tech in enumerate(technicians, 1):
                ws_tech.cell(row=i, column=1, value=f"{tech['name']} (ID: {tech['id']})")
            
            # Crear la validación que referencia la lista de técnicos
            dv = DataValidation(
                type="list",
                formula1=f'=Tecnicos!$A$1:$A${len(technicians)}',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle='Técnico no válido',
                error='Por favor seleccione un técnico de la lista'
            )
            ws.add_data_validation(dv)
            dv.add('A2:A1000')
            
            # 7. Guardar el archivo
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Plantilla Creada",
                f"Plantilla guardada en:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al crear la plantilla: {str(e)}"
            )
    
    def export_template_with_data(self):
        """Exporta los datos en formato de plantilla para importar en otro sistema"""
        try:
            if not hasattr(self, 'current_report') or not self.current_report.get('tasks'):
                QMessageBox.warning(self, "Error", "No hay datos para exportar")
                return
                
            # Obtener los datos actuales
            tasks = self.current_report['tasks']
            
            # Crear el DataFrame con las columnas de la plantilla
            data = []
            for task in tasks:
                # Obtener el nombre del técnico
                tech_name = ""
                for i in range(self.technician_combo.count()):
                    if self.technician_combo.itemData(i) == task['technician_id']:
                        tech_name = self.technician_combo.itemText(i)
                        break
                
                data.append({
                    'Técnico': tech_name,
                    'Cliente': task['client_name'],
                    'Tarea': task['task_description'],
                    'Presupuesto Total': task['budget_total'],
                    'Mano de Obra': task['labor_cost'],
                    'Presupuesto Materiales': task['material_cost'],
                    'Pago Seguro': task['insurance_payment'],
                    'Efectivo': task['cash_payment'],
                    'Número Pedido': task.get('order_number', ''),
                    'Gasto Material': task.get('material_expense', 0),
                    'Tipo de Pago': task.get('payment_type', 'EFECTIVO'),
                    'Estado': task.get('status', 'PENDIENTE'),
                    'Fecha (AAAA-MM-DD)': task['task_date']
                })
            
            df = pd.DataFrame(data)
            
            # Pedir ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Plantilla con Datos",
                f"tareas_exportadas_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Guardar en Excel
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Tareas', index=False)
                    
                    # Agregar hoja de técnicos para validación
                    technicians = self.db.get_technicians()
                    if technicians:
                        tech_data = [{'Técnico': f"{t['name']} (ID: {t['id']})"} for t in technicians]
                        pd.DataFrame(tech_data).to_excel(writer, sheet_name='Tecnicos', index=False)
                
                QMessageBox.information(
                    self, "Éxito", 
                    f"Datos exportados en formato de plantilla:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "Error al exportar", 
                f"Ocurrió un error al exportar los datos:\n{str(e)}"
            )
    
    def export_complete_data(self):
        """Exporta todos los datos de las tareas para análisis"""
        try:
            if not hasattr(self, 'current_report') or not self.current_report.get('tasks'):
                QMessageBox.warning(self, "Error", "No hay datos para exportar")
                return
                
            # Obtener los datos actuales
            tasks = self.current_report['tasks']
            df = pd.DataFrame(tasks)
            
            # Mapeo de nombres de columnas a español
            columnas_espanol = {
                'id': 'ID',
                'technician_id': 'ID Técnico',
                'client_name': 'Cliente',
                'task_description': 'Descripción',
                'task_date': 'Fecha',
                'budget_total': 'Presupuesto Total',
                'labor_cost': 'Costo Mano de Obra',
                'material_cost': 'Costo Materiales',
                'insurance_payment': 'Pago Seguro',
                'cash_payment': 'Pago Efectivo',
                'material_expense': 'Gasto Material',
                'profit': 'Ganancia',
                'pablo_share': 'Participación Técnico',
                'facu_share': 'Participación Socio',
                'iva': 'IVA',
                'payment_type': 'Tipo de Pago',
                'order_number': 'Número de Pedido',
                'status': 'Estado',
                'created_at': 'Fecha de Creación'
            }
            
            # Renombrar columnas
            df = df.rename(columns=columnas_espanol)
            
            # Formatear fechas
            if 'Fecha' in df.columns:
                df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%d')
            if 'Fecha de Creación' in df.columns:
                df['Fecha de Creación'] = pd.to_datetime(df['Fecha de Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Formatear montos monetarios
            for col in ['Presupuesto Total', 'Costo Mano de Obra', 'Costo Materiales', 
                       'Pago Seguro', 'Pago Efectivo', 'Gasto Material', 
                       'Ganancia', 'Participación Técnico', 'Participación Socio', 'IVA']:
                if col in df.columns:
                    df[col] = df[col].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "")
            
            # Pedir ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Datos Completos",
                f"datos_completos_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Guardar en Excel con formato
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Datos Completos')
                    
                    # Ajustar ancho de columnas
                    worksheet = writer.sheets['Datos Completos']
                    for idx, col in enumerate(df.columns, 1):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 30)
                
                QMessageBox.information(
                    self, "Éxito", 
                    f"Datos completos exportados:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "Error al exportar", 
                f"Ocurrió un error al exportar los datos completos:\n{str(e)}"
            )
    
    def export_technician_report(self):
        """Exporta un reporte de ganancias por técnico"""
        try:
            if not hasattr(self, 'current_report') or not self.current_report.get('tasks'):
                QMessageBox.warning(self, "Error", "No hay datos para generar el reporte")
                return
                
            # Obtener los datos actuales
            tasks = self.current_report['tasks']
            
            # Crear un diccionario para agrupar por técnico
            tech_data = {}
            
            for task in tasks:
                tech_id = task['technician_id']
                tech_name = ""
                
                # Buscar el nombre del técnico
                for i in range(self.technician_combo.count()):
                    if self.technician_combo.itemData(i) == tech_id:
                        tech_name = self.technician_combo.itemText(i)
                        break
                
                if tech_id not in tech_data:
                    tech_data[tech_id] = {
                        'Técnico': tech_name,
                        'Total Ventas': 0,
                        'Costo Mano de Obra': 0,
                        'Costo Materiales': 0,
                        'Ganancia Neta': 0,
                        'Cantidad de Trabajos': 0
                    }
                
                # Acumular valores
                tech_data[tech_id]['Total Ventas'] += task['budget_total']
                tech_data[tech_id]['Costo Mano de Obra'] += task['labor_cost']
                tech_data[tech_id]['Costo Materiales'] += task['material_cost']
                tech_data[tech_id]['Ganancia Neta'] += task['profit']
                tech_data[tech_id]['Cantidad de Trabajos'] += 1
            
            # Convertir a DataFrame
            df = pd.DataFrame(tech_data.values())
            
            # Ordenar por ganancia neta descendente
            df = df.sort_values('Ganancia Neta', ascending=False)
            
            # Agregar totales
            total_row = {
                'Técnico': 'TOTAL',
                'Total Ventas': df['Total Ventas'].sum(),
                'Costo Mano de Obra': df['Costo Mano de Obra'].sum(),
                'Costo Materiales': df['Costo Materiales'].sum(),
                'Ganancia Neta': df['Ganancia Neta'].sum(),
                'Cantidad de Trabajos': df['Cantidad de Trabajos'].sum()
            }
            df = df._append(total_row, ignore_index=True)
            
            # Pedir ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte por Técnico",
                f"reporte_tecnicos_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Guardar en Excel
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Resumen Técnicos', index=False)
                    
                    # Formatear la hoja
                    worksheet = writer.sheets['Resumen Técnicos']
                    for idx, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length
                    
                    # Resaltar la fila de totales
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=len(df) + 1, column=col)
                        cell.font = Font(bold=True)
                
                QMessageBox.information(
                    self, "Éxito", 
                    f"Reporte por técnico exportado:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "Error al exportar", 
                f"Ocurrió un error al generar el reporte por técnico:\n{str(e)}"
            )
    
    def export_facu_report(self):
        """Exporta el reporte de la participación de Facu (30%)"""
        try:
            if not hasattr(self, 'current_report') or not self.current_report.get('tasks'):
                QMessageBox.warning(self, "Error", "No hay datos para generar el reporte")
                return
                
            # Obtener los datos actuales
            tasks = self.current_report['tasks']
            
            # Calcular totales
            total_ganancia = sum(task['profit'] for task in tasks)
            total_facu = sum(task.get('facu_share', 0) for task in tasks)
            total_tecnico = sum(task.get('pablo_share', 0) for task in tasks)
            
            # Crear DataFrame con el resumen
            data = [
                {
                    'Concepto': 'Ganancia Total',
                    'Monto': total_ganancia,
                    'Porcentaje': '100%'
                },
                {
                    'Concepto': 'Participación de Socio (30%)',
                    'Monto': total_facu,
                    'Porcentaje': '30%'
                },
                {
                    'Concepto': 'Participación de Técnico (70%)',
                    'Monto': total_tecnico,
                    'Porcentaje': '70%'
                }
            ]
            
            df = pd.DataFrame(data)
            
            # Agregar detalle por tarea
            detalle = []
            for task in tasks:
                if task.get('facu_share', 0) > 0:
                    detalle.append({
                        'Fecha': task['task_date'],
                        'Cliente': task['client_name'],
                        'Descripción': task['task_description'],
                        'Ganancia Total': task['profit'],
                        'Socio (30%)': task.get('facu_share', 0),
                        'Técnico (70%)': task.get('pablo_share', 0)
                    })
            
            df_detalle = pd.DataFrame(detalle)
            
            # Agregar fila de totales al detalle
            if not df_detalle.empty:
                total_row = {
                    'Fecha': 'TOTAL',
                    'Cliente': '',
                    'Descripción': '',
                    'Ganancia Total': df_detalle['Ganancia Total'].sum(),
                    'Socio (30%)': df_detalle['Socio (30%)'].sum(),
                    'Técnico (70%)': df_detalle['Técnico (70%)'].sum()
                }
                df_detalle = df_detalle._append(total_row, ignore_index=True)
            
            # Pedir ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte de Socio",
                f"reporte_socio_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Archivos Excel (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Guardar en Excel
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Hoja de resumen
                    df.to_excel(writer, sheet_name='Resumen', index=False)
                    
                    # Formatear hoja de resumen
                    worksheet = writer.sheets['Resumen']
                    for idx, col in enumerate(df.columns, 1):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        ) + 2
                        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 30)
                    
                    # Resaltar totales en negrita
                    for row in range(2, 5):  # Filas 2 a 4 (los totales)
                        for col in range(1, 4):  # Columnas A a C
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = Font(bold=True)
                    
                    # Hoja de detalle si hay datos
                    if not df_detalle.empty:
                        df_detalle.to_excel(writer, sheet_name='Detalle', index=False)
                        
                        # Formatear hoja de detalle
                        worksheet = writer.sheets['Detalle']
                        for idx, col in enumerate(df_detalle.columns, 1):
                            max_length = max(
                                df_detalle[col].astype(str).apply(len).max(),
                                len(str(col))
                            ) + 2
                            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 30)
                        
                        # Formato de moneda para las columnas numéricas
                        money_columns = ['Ganancia Total', 'Socio (30%)', 'Técnico (70%)']
                        for col_num, col_name in enumerate(df_detalle.columns, 1):
                            if col_name in money_columns:
                                for row in range(2, len(df_detalle) + 2):  # +2 porque Excel es 1-based y tiene encabezado
                                    cell = worksheet.cell(row=row, column=col_num)
                                    if row == len(df_detalle) + 1:  # Fila de totales
                                        cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'
                                        cell.font = Font(bold=True)
                                    else:
                                        cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'
                
                QMessageBox.information(
                    self, "Éxito", 
                    f"Reporte de Facu exportado:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, "Error al exportar", 
                f"Ocurrió un error al generar el reporte de Facu:\n{str(e)}"
            )
